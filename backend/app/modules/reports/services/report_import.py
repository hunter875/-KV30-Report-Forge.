from __future__ import annotations

import csv
import io
import unicodedata
from datetime import datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.modules.reports.models import Report, ReportRow, RowType


COLUMN_KEYS: dict[RowType, list[str]] = {
    RowType.STATISTICS: [
        "ngay",
        "thang",
        "vu_chay_thong_ke",
        "sclq_den_pccc_cnch",
        "chi_vien",
        "cnch",
        "dinh_ky_nhom_i",
        "dinh_ky_nhom_ii",
        "dot_xuat_nhom_i",
        "dot_xuat_nhom_ii",
        "huong_dan",
        "kien_nghi",
        "xu_phat",
        "tien_phat_trieu_dong",
        "dinh_chi",
        "phuc_hoi",
        "tin_bai",
        "phong_su",
        "so_lop_tuyen_truyen",
        "so_nguoi_tham_du_tuyen_truyen",
        "so_khuyen_cao_to_roi_da_phat",
        "so_lop_huan_luyen",
        "so_nguoi_tham_du_huan_luyen",
        "tong_so_lop",
        "tong_so_nguoi_tham_du",
        "pc06_so_pa_xay_dung_va_phe_duyet",
        "pc06_so_pa_duoc_thuc_tap",
        "pc08_so_pa_xay_dung_va_phe_duyet",
        "pc08_so_pa_duoc_thuc_tap",
        "pc09_so_pa_xay_dung_va_phe_duyet",
        "pc09_so_pa_duoc_thuc_tap",
        "pc07_so_pa_xay_dung_va_phe_duyet",
        "pc07_so_pa_duoc_thuc_tap",
        "ghi_chu",
    ],
    RowType.CNCH_EVENT: [
        "loai_hinh_cnch",
        "ngay_xay_ra",
        "thoi_gian",
        "dia_diem",
        "dia_chi",
        "chi_huy_cnch",
        "thiet_hai_ve_nguoi",
        "so_nguoi_cuu_duoc",
    ],
    RowType.SCLQ: [
        "vu_chay_ngay",
        "dia_diem",
        "nguyen_nhan",
        "thiet_hai",
        "chi_huy_chua_chay",
        "ghi_chu",
    ],
    RowType.TONG_VU_CHAY: [
        "ngay_xay_ra",
        "vu_chay",
        "thoi_gian",
        "dia_diem",
        "phan_loai",
        "nguyen_nhan",
        "thiet_hai_ve_nguoi",
        "thiet_hai_tai_san",
        "tai_san_cuu_chua",
        "thoi_gian_toi_dam_chay",
        "thoi_gian_khong_che",
        "thoi_gian_dap_tat_hoan_toan",
        "so_luong_xe",
        "chi_huy_chua_chay",
        "ghi_chu",
    ],
}


XLSX_LAYOUTS: dict[RowType, dict[str, Any]] = {
    RowType.STATISTICS: {
        "sheet_names": ["BC NGÀY"],
        "start_row": 4,
        "min_col": 1,
        "max_col": 34,
    },
    RowType.CNCH_EVENT: {
        "sheet_names": ["CNCH"],
        "start_row": 3,
        "min_col": 1,
        "max_col": 9,
    },
    RowType.SCLQ: {
        "sheet_names": ["SCLQ ĐẾN PCCC&CNCH", "SCLQ"],
        "start_row": 3,
        "min_col": 1,
        "max_col": 7,
    },
    RowType.TONG_VU_CHAY: {
        "sheet_names": ["VỤ CHÁY THỐNG KÊ"],
        "start_row": 3,
        "min_col": 1,
        "max_col": 16,
    },
}


class ReportImportService:
    def __init__(self, db: Session):
        self.db = db

    def import_file(
        self,
        report_id: str,
        row_type: str,
        filename: str,
        content: bytes,
        replace: bool = False,
    ) -> dict[str, Any]:
        try:
            report = self.db.query(Report).filter_by(id=report_id).with_for_update().first()
            if not report:
                raise ValueError("Report not found")

            parsed_row_type = self._parse_row_type(row_type)
            if parsed_row_type not in COLUMN_KEYS:
                raise ValueError(f"Unsupported import row_type: {row_type}")

            rows = self._read_rows(parsed_row_type, filename, content)
            payloads = self._map_rows(parsed_row_type, rows)

            if replace:
                (
                    self.db.query(ReportRow)
                    .filter(ReportRow.report_id == report.id, ReportRow.row_type == parsed_row_type)
                    .delete(synchronize_session=False)
                )
                next_stt = 1
            else:
                max_stt = (
                    self.db.query(ReportRow.stt)
                    .filter(ReportRow.report_id == report.id, ReportRow.row_type == parsed_row_type)
                    .order_by(ReportRow.stt.desc().nullslast())
                    .first()
                )
                next_stt = (max_stt[0] if max_stt and max_stt[0] else 0) + 1

            for index, payload in enumerate(payloads, start=next_stt):
                payload["stt"] = index
                self.db.add(
                    ReportRow(
                        report_id=report.id,
                        row_type=parsed_row_type,
                        stt=index,
                        payload=payload,
                    )
                )

            report.version += 1
            self.db.commit()
            self.db.refresh(report)

            return {
                "status": "ok",
                "row_type": parsed_row_type.value,
                "imported": len(payloads),
                "version": report.version,
            }
        except Exception as exc:
            self.db.rollback()
            raise ValueError(f"Import thất bại: {str(exc)}") from exc

    def import_workbook(
        self,
        report_id: str,
        filename: str,
        content: bytes,
        replace: bool = True,
    ) -> dict[str, Any]:
        try:
            report = self.db.query(Report).filter_by(id=report_id).with_for_update().first()
            if not report:
                raise ValueError("Report not found")

            if not filename.lower().endswith(".xlsx"):
                raise ValueError("Import 3 bảng chỉ hỗ trợ file .xlsx gốc.")

            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
            targets = [
                RowType.STATISTICS,
                RowType.CNCH_EVENT,
                RowType.SCLQ,
            ]
            counts: dict[str, int] = {}

            for row_type in targets:
                rows = self._read_xlsx_rows(workbook, row_type)
                payloads = self._map_rows(row_type, rows)

                if replace:
                    (
                        self.db.query(ReportRow)
                        .filter(ReportRow.report_id == report.id, ReportRow.row_type == row_type)
                        .delete(synchronize_session=False)
                    )
                    next_stt = 1
                else:
                    max_stt = (
                        self.db.query(ReportRow.stt)
                        .filter(ReportRow.report_id == report.id, ReportRow.row_type == row_type)
                        .order_by(ReportRow.stt.desc().nullslast())
                        .first()
                    )
                    next_stt = (max_stt[0] if max_stt and max_stt[0] else 0) + 1

                for index, payload in enumerate(payloads, start=next_stt):
                    payload["stt"] = index
                    self.db.add(
                        ReportRow(
                            report_id=report.id,
                            row_type=row_type,
                            stt=index,
                            payload=payload,
                        )
                    )

                counts[row_type.value] = len(payloads)

            report.version += 1
            self.db.commit()
            self.db.refresh(report)

            return {
                "status": "ok",
                "imported": counts,
                "version": report.version,
            }
        except Exception as exc:
            self.db.rollback()
            raise ValueError(f"Import workbook thất bại: {str(exc)}") from exc

    def preview_file(
        self,
        row_type: str,
        filename: str,
        content: bytes,
    ) -> dict[str, Any]:
        """Parse one import target without writing to the database."""
        parsed_row_type = self._parse_row_type(row_type)
        if parsed_row_type not in COLUMN_KEYS:
            raise ValueError(f"Unsupported import row_type: {row_type}")

        rows = self._read_rows(parsed_row_type, filename, content)
        payloads = self._map_rows(parsed_row_type, rows)
        return {
            "status": "ok",
            "mode": "single",
            "can_import": len(payloads) > 0 and not self._date_diagnostics(parsed_row_type, payloads)[0],
            "sheets": [
                self._build_preview_item(
                    row_type=parsed_row_type,
                    filename=filename,
                    raw_rows=len(rows),
                    payloads=payloads,
                    sheet_name=self._preview_sheet_name(parsed_row_type, filename, content),
                )
            ],
        }

    def preview_workbook(self, filename: str, content: bytes) -> dict[str, Any]:
        """Parse the three source sheets without writing to the database."""
        if not filename.lower().endswith(".xlsx"):
            raise ValueError("Preview 3 bảng chỉ hỗ trợ file .xlsx gốc.")

        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
        targets = [RowType.STATISTICS, RowType.CNCH_EVENT, RowType.SCLQ]
        previews = []
        can_import = True

        for row_type in targets:
            try:
                sheet = self._find_sheet(workbook, XLSX_LAYOUTS[row_type]["sheet_names"])
                rows = self._read_xlsx_rows(workbook, row_type)
                payloads = self._map_rows(row_type, rows)
                item = self._build_preview_item(
                    row_type=row_type,
                    filename=filename,
                    raw_rows=len(rows),
                    payloads=payloads,
                    sheet_name=sheet.title,
                    header_warnings=self._header_warnings(sheet, row_type),
                )
                previews.append(item)
                can_import = can_import and item["row_count"] > 0 and not item["errors"] and not item["date_errors"]
            except Exception as exc:
                previews.append({
                    "row_type": row_type.value,
                    "label": self._row_type_label(row_type),
                    "sheet_name": "",
                    "row_count": 0,
                    "raw_rows": 0,
                    "date_errors": [],
                    "warnings": [],
                    "errors": [str(exc)],
                    "sample_rows": [],
                    "sample_dates": [],
                })
                can_import = False

        return {
            "status": "ok",
            "mode": "workbook",
            "can_import": can_import,
            "sheets": previews,
        }

    @staticmethod
    def _parse_row_type(row_type: str) -> RowType:
        normalized = row_type.lower().replace(" ", "_")
        try:
            return RowType(normalized)
        except ValueError as exc:
            raise ValueError(f"Unsupported row_type: {row_type}") from exc

    @staticmethod
    def _read_rows(row_type: RowType, filename: str, content: bytes) -> list[list[Any]]:
        lower = filename.lower()
        if lower.endswith(".csv"):
            text = content.decode("utf-8-sig")
            return [row for row in csv.reader(io.StringIO(text))]

        if lower.endswith(".tsv"):
            text = content.decode("utf-8-sig")
            return [row for row in csv.reader(io.StringIO(text), delimiter="\t")]

        if lower.endswith(".xls"):
            raise ValueError("File .xls cũ chưa được hỗ trợ. Hãy xuất Google Sheet/Excel thành .xlsx hoặc .csv.")

        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
        return ReportImportService._read_xlsx_rows(workbook, row_type)

    @staticmethod
    def _read_xlsx_rows(workbook: Any, row_type: RowType) -> list[list[Any]]:
        """Read the KV30 workbook using fixed sheet layouts.

        The source workbook uses merged, multi-level headers. Header inference is
        the wrong tool here, so imports are mapped by known sheet names and cell
        coordinates instead.
        """
        layout = XLSX_LAYOUTS.get(row_type)
        if not layout:
            raise ValueError(f"Unsupported Excel import row_type: {row_type.value}")

        sheet = ReportImportService._find_sheet(workbook, layout["sheet_names"])
        rows: list[list[Any]] = []

        for row in sheet.iter_rows(
            min_row=layout["start_row"],
            min_col=layout["min_col"],
            max_col=layout["max_col"],
            values_only=True,
        ):
            values = list(row)
            if ReportImportService._is_empty_source_row(row_type, values):
                continue
            rows.append(values)

        if not rows:
            expected = ", ".join(layout["sheet_names"])
            raise ValueError(f"Không đọc được dữ liệu trong sheet {expected}")

        return rows

    @staticmethod
    def _find_sheet(workbook: Any, candidates: list[str]) -> Any:
        normalized_candidates = [ReportImportService._normalize_text(name) for name in candidates]

        for candidate in normalized_candidates:
            for sheet_name in workbook.sheetnames:
                if ReportImportService._normalize_text(sheet_name) == candidate:
                    return workbook[sheet_name]

        for candidate in normalized_candidates:
            for sheet_name in workbook.sheetnames:
                normalized_sheet = ReportImportService._normalize_text(sheet_name)
                if normalized_sheet.startswith(candidate) or candidate in normalized_sheet:
                    return workbook[sheet_name]

        available = ", ".join(workbook.sheetnames)
        expected = ", ".join(candidates)
        raise ValueError(f"Không tìm thấy sheet {expected}. Các sheet hiện có: {available}")

    @staticmethod
    def _normalize_text(value: Any) -> str:
        text = str(value or "").strip().lower()
        text = unicodedata.normalize("NFD", text)
        text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
        text = text.replace("&", " ")
        return " ".join(text.split())

    @staticmethod
    def _is_empty_source_row(row_type: RowType, values: list[Any]) -> bool:
        if not ReportImportService._trim_row(values):
            return True

        if row_type == RowType.STATISTICS:
            ngay = values[0] if len(values) > 0 else None
            thang = values[1] if len(values) > 1 else None
            return ngay in (None, "") or thang in (None, "")

        # Detail sheets have STT in column A. Ignore formatting-only blank rows.
        return all(value in (None, "") for value in values[1:])

    @classmethod
    def _map_rows(cls, row_type: RowType, rows: list[list[Any]]) -> list[dict[str, Any]]:
        column_keys = COLUMN_KEYS[row_type]
        data_rows = cls._data_rows(rows)
        result = []

        for row in data_rows:
            values = cls._trim_row(row)
            if not values:
                continue

            if row_type != RowType.STATISTICS and not cls._looks_like_stt(values[0]):
                continue

            if cls._looks_like_stt(values[0]) and row_type != RowType.STATISTICS:
                values = values[1:]

            payload = {}
            for index, key in enumerate(column_keys):
                value = values[index] if index < len(values) else ""
                payload[key] = cls._clean_value(value)

            if any(value not in (None, "") for value in payload.values()):
                result.append(payload)

        return result

    @staticmethod
    def _data_rows(rows: list[list[Any]]) -> list[list[Any]]:
        start_index = 0
        for index, row in enumerate(rows[:20]):
            normalized_cells = [ReportImportService._normalize_text(cell) for cell in row[:4]]
            normalized = " ".join(normalized_cells)
            if "stt" in normalized or "noi dung" in normalized:
                start_index = index + 1
        return rows[start_index:]

    @staticmethod
    def _trim_row(row: list[Any]) -> list[Any]:
        values = list(row)
        while values and values[-1] in (None, ""):
            values.pop()
        return values

    @staticmethod
    def _looks_like_stt(value: Any) -> bool:
        if isinstance(value, int):
            return True
        if isinstance(value, float) and value.is_integer():
            return True
        return str(value or "").strip().isdigit()

    @classmethod
    def _build_preview_item(
        cls,
        row_type: RowType,
        filename: str,
        raw_rows: int,
        payloads: list[dict[str, Any]],
        sheet_name: str,
        header_warnings: list[str] | None = None,
    ) -> dict[str, Any]:
        date_errors, sample_dates = cls._date_diagnostics(row_type, payloads)
        warnings = list(header_warnings or [])
        if raw_rows != len(payloads):
            warnings.append(f"Đọc {raw_rows} dòng nguồn, map được {len(payloads)} dòng dữ liệu.")
        if not payloads:
            warnings.append("Không có dòng dữ liệu hợp lệ để import.")

        return {
            "row_type": row_type.value,
            "label": cls._row_type_label(row_type),
            "sheet_name": sheet_name or filename,
            "row_count": len(payloads),
            "raw_rows": raw_rows,
            "date_errors": date_errors,
            "warnings": warnings,
            "errors": [],
            "sample_rows": payloads[:5],
            "sample_dates": sample_dates,
        }

    @staticmethod
    def _preview_sheet_name(row_type: RowType, filename: str, content: bytes) -> str:
        if not filename.lower().endswith(".xlsx"):
            return filename
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
            sheet = ReportImportService._find_sheet(workbook, XLSX_LAYOUTS[row_type]["sheet_names"])
            return sheet.title
        except Exception:
            return filename

    @staticmethod
    def _header_warnings(sheet: Any, row_type: RowType) -> list[str]:
        header_text = ReportImportService._normalize_text(
            " ".join(
                str(cell.value or "")
                for row in sheet.iter_rows(min_row=1, max_row=3)
                for cell in row
            )
        )
        expected_tokens = {
            RowType.STATISTICS: ("ngay", "thang"),
            RowType.CNCH_EVENT: ("cnch", "dia diem"),
            RowType.SCLQ: ("pccc", "thiet hai"),
        }.get(row_type, ())
        missing = [token for token in expected_tokens if token not in header_text]
        if not missing:
            return []
        return [f"Header sheet {sheet.title} thiếu token kỳ vọng: {', '.join(missing)}. Vẫn map theo tọa độ cố định."]

    @staticmethod
    def _date_diagnostics(row_type: RowType, payloads: list[dict[str, Any]]) -> tuple[list[str], list[str]]:
        date_keys = {
            RowType.STATISTICS: ("ngay", "thang"),
            RowType.CNCH_EVENT: ("ngay_xay_ra",),
            RowType.SCLQ: ("vu_chay_ngay",),
        }
        errors: list[str] = []
        sample_dates: list[str] = []

        for index, payload in enumerate(payloads, start=1):
            if row_type == RowType.STATISTICS:
                day = ReportImportService._to_int(payload.get("ngay"))
                month = ReportImportService._to_int(payload.get("thang"))
                if not (1 <= day <= 31 and 1 <= month <= 12):
                    errors.append(f"Dòng {index}: ngày/tháng không hợp lệ ({payload.get('ngay')}/{payload.get('thang')})")
                    continue
                label = f"{day:02d}/{month:02d}"
            else:
                key = date_keys.get(row_type, ("",))[0]
                value = payload.get(key)
                if not ReportImportService._parse_date(value):
                    errors.append(f"Dòng {index}: ngày không hợp lệ ({value})")
                    continue
                label = str(value)

            if label not in sample_dates and len(sample_dates) < 8:
                sample_dates.append(label)

        return errors[:20], sample_dates

    @staticmethod
    def _parse_date(value: Any):
        if value in (None, ""):
            return None
        if hasattr(value, "date") and not isinstance(value, str):
            try:
                return value.date()
            except Exception:
                return value
        text = str(value).strip()
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    @staticmethod
    def _to_int(value: Any) -> int:
        if value in (None, ""):
            return 0
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, (int, float)):
            return int(value)
        try:
            return int(float(str(value).replace(",", ".").strip()))
        except ValueError:
            return 0

    @staticmethod
    def _row_type_label(row_type: RowType) -> str:
        return {
            RowType.STATISTICS: "BC NGÀY",
            RowType.CNCH_EVENT: "CNCH",
            RowType.SCLQ: "SCLQ",
            RowType.TONG_VU_CHAY: "TỔNG VỤ CHÁY",
        }.get(row_type, row_type.value)

    @staticmethod
    def _clean_value(value: Any) -> Any:
        if value is None:
            return ""
        if hasattr(value, "strftime"):
            return value.strftime("%d/%m/%Y")
        return value
